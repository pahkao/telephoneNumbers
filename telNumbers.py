import openpyxl as xl
import numpy as np
import re
from openpyxl import load_workbook, Workbook
from os.path import join, abspath


class NotAllData(Exception):
    pass


data_path = join('.', '1.xlsx')  # текущая директория + имя файла
data_path = abspath(data_path)

wb = load_workbook(filename=data_path, data_only=True, read_only=True)  # читаем файл
ws = wb.active

l1 = list()
l2 = list()
l3 = list()
l4 = list()

for row in ws.values:   # добавляем все значения в ячейках активного листа (в excel) в лист 'l1'
    for value in row:
        if value:
            l1.append(value)

for i in l1:    # ищем номера в листе 'l1' типа: +X(XXX)XXX-XX-XX
    tel = str(i)
    tel2 = (re.findall('\+([7-8]\([0-9]+\)[0-9]+\-[0-9]+\-[0-9]+)', tel))
    if tel2:
        l2.append(tel2)

for x in l2:    # добавляем в лист l3 только уникальные номера из листа l2
    for y in x:
        if y not in l3:
            l3.append(y)

for x in l3:    # убираем лишние символы из номеров ')', '(', '+', '-'
    tel3 = re.sub('[^\d+]', '', x)
    l4.append(tel3)


wb1 = Workbook()    # создаем новую книгу, добавляем номера в 1ый столбец, сохраняем

ws1 = wb1.active
ws1.title = "telNumbers"

r = 1
for n in l4:
    ws1.cell(row=r, column=1).value = n
    r += 1

wb1.save('numbers.xlsx')
